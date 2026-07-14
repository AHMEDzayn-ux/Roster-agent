"""Excel export/import for manager roster edits (spec §2.7): export the
generated roster to .xlsx, let the manager edit it locally, re-upload, and
re-validate against hard constraints before applying.

Hard constraints (always rejected, nothing applied): double-booking,
unknown/mismatched agent-shift-skill references, coverage shortfalls, and
scheduling work on a day covered by an already-*approved* leave request
("ignores a locked leave day" — spec's own example of a hard violation).

Soft: if the edit flips a solver-decided off-day/leave/shift-change/overtime
outcome (in either direction) it requires a stated reason (the "supervisor
sign-off" safeguard) and is written to the audit trail.
"""
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from app.crud.weekly_request import LEAVE_REQUEST_TYPES
from app.domain.leave import leave_dates
from app.domain.shift_coverage import shift_covers_slot
from app.models.agent import Agent, AgentSkill
from app.models.audit import AuditLog
from app.models.coverage import CoverageRequirement
from app.models.enums import AssignmentSource, RequestStatus, RequestType, WeeklyCycleStatus
from app.models.roster import Roster, RosterAssignment
from app.models.shift import ShiftTemplate
from app.models.skill import Skill
from app.models.weekly_cycle import WeeklyCycle
from app.models.weekly_request import WeeklyRequest
from app.services.leave_balance_sync import LeaveBalanceSyncError, sync_leave_balance

EXPORT_HEADERS = ["agent_id", "agent_name", "date", "shift_id", "shift_name", "skill_id", "skill_name"]
IMPORT_REQUIRED_COLUMNS = {"agent_id", "date", "shift_id", "skill_id"}

# Machine-readable sheet the re-upload/import path parses back. The first sheet
# ("Roster") is the human-friendly grid; this one carries the raw ids.
DATA_SHEET_NAME = "Data"

# Light fills per shift so the grid reads like the Roster Workspace table
# (kept close to the app's shift colours). Overnight uses a dark fill + white text.
_SHIFT_FILLS = {
    "Morning": "E0F2FE",
    "Early Morning": "FEF3C7",
    "Mid Morning": "CCFBF1",
    "Afternoon": "EDE9FE",
    "Evening": "FFE4E6",
    "Overnight": "312E81",
}
_DEFAULT_SHIFT_FILL = "F1F5F9"
_WHITE_FONT_SHIFTS = {"Overnight"}


def _compact_12h(t) -> str:
    """Bare 12-hour hour, matching the workspace: 12:00-21:00 -> '12-9'."""
    hour = t.hour % 12
    return str(12 if hour == 0 else hour)


class RosterImportError(Exception):
    def __init__(self, detail: str, violations: list[str] | None = None, status_code: int = 422):
        self.detail = detail
        self.violations = violations or []
        self.status_code = status_code
        super().__init__(detail)


@dataclass
class ParsedAssignmentRow:
    agent_id: int
    date: date
    shift_id: int
    skill_id: int


@dataclass
class ImportResult:
    roster: Roster
    overridden_requests: list[str]


def export_roster_workbook(db: Session, roster: Roster) -> Workbook:
    assignments = (
        db.query(RosterAssignment)
        .options(
            joinedload(RosterAssignment.agent),
            joinedload(RosterAssignment.shift),
            joinedload(RosterAssignment.skill_covered),
        )
        .filter(RosterAssignment.roster_id == roster.id)
        .order_by(RosterAssignment.date, RosterAssignment.agent_id)
        .all()
    )
    wb = Workbook()
    _write_grid_sheet(wb.active, assignments)
    _write_data_sheet(wb.create_sheet(DATA_SHEET_NAME), assignments)
    wb.active = 0  # open on the human-friendly grid
    return wb


def _write_grid_sheet(ws, assignments: list[RosterAssignment]) -> None:
    """The Roster Workspace layout: one row per agent, one column per day
    (Mon–Sun), each cell the compact shift time (e.g. '12-9') colour-coded by
    shift, 'OFF' where the agent isn't scheduled."""
    ws.title = "Roster"

    dates = sorted({a.date for a in assignments})
    agent_names = {a.agent_id: a.agent.name for a in assignments}
    agents = sorted(agent_names.items(), key=lambda kv: kv[1])
    by_agent_date = {(a.agent_id, a.date): a for a in assignments}

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="F1F5F9")
    center = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value="Agent").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    for col, d in enumerate(dates, start=2):
        cell = ws.cell(row=1, column=col, value=f"{d.strftime('%A')}\n{d.isoformat()}")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r, (agent_id, name) in enumerate(agents, start=2):
        ws.cell(row=r, column=1, value=name)
        for col, d in enumerate(dates, start=2):
            cell = ws.cell(row=r, column=col)
            a = by_agent_date.get((agent_id, d))
            if a is None:
                cell.value = "OFF"
                cell.font = Font(color="94A3B8")
            else:
                cell.value = f"{_compact_12h(a.shift.start_time)}-{_compact_12h(a.shift.end_time)}"
                cell.fill = PatternFill("solid", fgColor=_SHIFT_FILLS.get(a.shift.name, _DEFAULT_SHIFT_FILL))
                if a.shift.name in _WHITE_FONT_SHIFTS:
                    cell.font = Font(color="FFFFFF")
                cell.comment = None
            cell.alignment = center

    ws.column_dimensions["A"].width = 24
    for col in range(2, len(dates) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = "B2"


def _write_data_sheet(ws, assignments: list[RosterAssignment]) -> None:
    """Machine-readable long format used by the re-upload/import path."""
    ws.append(EXPORT_HEADERS)
    for a in assignments:
        ws.append(
            [
                a.agent_id,
                a.agent.name,
                a.date.isoformat(),
                a.shift_id,
                a.shift.name,
                a.skill_covered_id,
                a.skill_covered.name,
            ]
        )


def parse_import_file(file_bytes: bytes) -> list[ParsedAssignmentRow]:
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise RosterImportError(
            "Could not read the uploaded file as an Excel (.xlsx) workbook", status_code=400
        ) from exc

    # Prefer the machine-readable "Data" sheet (the pretty grid is the first
    # sheet now); fall back to the active sheet for older flat-format uploads.
    ws = wb[DATA_SHEET_NAME] if DATA_SHEET_NAME in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RosterImportError("The uploaded file is empty", status_code=400)

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    col_index = {name: i for i, name in enumerate(header)}
    missing = IMPORT_REQUIRED_COLUMNS - set(col_index.keys())
    if missing:
        raise RosterImportError(f"Missing required column(s): {', '.join(sorted(missing))}", status_code=400)

    parsed: list[ParsedAssignmentRow] = []
    errors: list[str] = []
    for row_num, row in enumerate(rows[1:], start=2):
        if row is None or all(v is None for v in row):
            continue
        try:
            agent_id = int(row[col_index["agent_id"]])
            shift_id = int(row[col_index["shift_id"]])
            skill_id = int(row[col_index["skill_id"]])
            raw_date = row[col_index["date"]]
            if isinstance(raw_date, datetime):
                parsed_date = raw_date.date()
            elif isinstance(raw_date, date):
                parsed_date = raw_date
            else:
                parsed_date = date.fromisoformat(str(raw_date).strip())
            parsed.append(ParsedAssignmentRow(agent_id=agent_id, date=parsed_date, shift_id=shift_id, skill_id=skill_id))
        except (TypeError, ValueError) as exc:
            errors.append(f"Row {row_num}: {exc}")

    if errors:
        raise RosterImportError("Could not parse some rows", violations=errors, status_code=400)
    return parsed


def _validate_hard_constraints(db: Session, cycle: WeeklyCycle, rows: list[ParsedAssignmentRow]) -> list[str]:
    violations: list[str] = []

    seen = set()
    for r in rows:
        key = (r.agent_id, r.date)
        if key in seen:
            violations.append(f"Agent {r.agent_id} is double-booked on {r.date}")
        seen.add(key)

    agent_ids = {r.agent_id for r in rows}
    shift_ids = {r.shift_id for r in rows}
    skill_ids = {r.skill_id for r in rows}

    agents = {a.id: a for a in db.query(Agent).filter(Agent.id.in_(agent_ids)).all()} if agent_ids else {}
    shifts = (
        {s.id: s for s in db.query(ShiftTemplate).filter(ShiftTemplate.id.in_(shift_ids)).all()} if shift_ids else {}
    )
    skills = {s.id: s for s in db.query(Skill).filter(Skill.id.in_(skill_ids)).all()} if skill_ids else {}
    agent_skill_ids: dict[int, set[int]] = {}
    if agent_ids:
        for link in db.query(AgentSkill).filter(AgentSkill.agent_id.in_(agent_ids)).all():
            agent_skill_ids.setdefault(link.agent_id, set()).add(link.skill_id)

    for r in rows:
        if r.agent_id not in agents:
            violations.append(f"Agent {r.agent_id} does not exist (row for {r.date})")
        if r.shift_id not in shifts:
            violations.append(f"Shift template {r.shift_id} does not exist (row for agent {r.agent_id}, {r.date})")
        if r.skill_id not in skills:
            violations.append(f"Skill {r.skill_id} does not exist (row for agent {r.agent_id}, {r.date})")
        if r.agent_id in agents and r.skill_id in skills and r.skill_id not in agent_skill_ids.get(r.agent_id, set()):
            violations.append(f"Agent {r.agent_id} does not hold skill {r.skill_id} (row for {r.date})")

    # Only check coverage once references are all valid (shift lookups below assume this)
    if not violations:
        coverage_reqs = db.query(CoverageRequirement).all()
        week_dates = [cycle.week_start_date + timedelta(days=i) for i in range(7)]
        for req in coverage_reqs:
            matching_dates = [wd for wd in week_dates if wd.weekday() == req.day_of_week]
            for d in matching_dates:
                count = 0
                for r in rows:
                    if r.date != d or r.skill_id != req.skill_id:
                        continue
                    shift = shifts[r.shift_id]
                    if shift_covers_slot(shift.start_time, shift.end_time, req.time_slot_start, req.time_slot_end):
                        count += 1
                if count < req.min_agents_required:
                    violations.append(
                        f"Coverage shortfall on {d} for skill {req.skill_id} "
                        f"({req.time_slot_start}-{req.time_slot_end}): needs {req.min_agents_required}, has {count}"
                    )

    approved_leaves = (
        db.query(WeeklyRequest)
        .filter(
            WeeklyRequest.week_cycle_id == cycle.id,
            WeeklyRequest.status == RequestStatus.approved,
            WeeklyRequest.request_type.in_(LEAVE_REQUEST_TYPES),
        )
        .all()
    )
    worked_dates_by_agent: dict[int, set[date]] = {}
    for r in rows:
        worked_dates_by_agent.setdefault(r.agent_id, set()).add(r.date)
    for req in approved_leaves:
        for d in leave_dates(req.request_type.value, req.requested_start_date, req.requested_end_date):
            if d in worked_dates_by_agent.get(req.agent_id, set()):
                violations.append(
                    f"Agent {req.agent_id} has approved leave on {d}; the upload schedules them to work that day"
                )

    return violations


def _compute_flips(db: Session, cycle: WeeklyCycle, rows: list[ParsedAssignmentRow]) -> list[tuple[WeeklyRequest, bool]]:
    worked_dates_by_agent: dict[int, set[date]] = {}
    shift_by_agent_date: dict[tuple[int, date], int] = {}
    for r in rows:
        worked_dates_by_agent.setdefault(r.agent_id, set()).add(r.date)
        shift_by_agent_date[(r.agent_id, r.date)] = r.shift_id

    candidates = (
        db.query(WeeklyRequest)
        .filter(
            WeeklyRequest.week_cycle_id == cycle.id,
            WeeklyRequest.status.in_([RequestStatus.approved, RequestStatus.denied]),
            WeeklyRequest.request_type != RequestType.other,
        )
        .all()
    )

    flips = []
    for req in candidates:
        rtype = req.request_type
        if rtype in LEAVE_REQUEST_TYPES:
            dates = leave_dates(rtype.value, req.requested_start_date, req.requested_end_date)
            proposed_honored = all(d not in worked_dates_by_agent.get(req.agent_id, set()) for d in dates)
        elif rtype == RequestType.off_day:
            proposed_honored = req.requested_start_date not in worked_dates_by_agent.get(req.agent_id, set())
        elif rtype == RequestType.overtime:
            proposed_honored = req.requested_start_date in worked_dates_by_agent.get(req.agent_id, set())
        elif rtype == RequestType.shift_change:
            proposed_shift = shift_by_agent_date.get((req.agent_id, req.requested_start_date))
            proposed_honored = proposed_shift == req.requested_shift_id
        else:
            continue

        currently_honored = req.status == RequestStatus.approved
        if proposed_honored != currently_honored:
            flips.append((req, proposed_honored))

    return flips


def apply_proposed_roster(
    db: Session,
    roster: Roster,
    cycle: WeeklyCycle,
    rows: list[ParsedAssignmentRow],
    reason: str | None,
    actor_id: int,
    action_type: str,
    require_reason: bool,
) -> ImportResult:
    """Shared core for both Excel re-upload and the single-assignment manual
    override endpoint: validate hard constraints, detect soft-constraint
    flips against currently-decided request outcomes, apply if clean (or a
    reason was given), and write the audit trail."""
    if cycle.status == WeeklyCycleStatus.locked:
        raise RosterImportError("This weekly cycle is locked; the roster can no longer be edited", status_code=400)

    violations = _validate_hard_constraints(db, cycle, rows)
    if violations:
        raise RosterImportError(
            "The uploaded roster violates hard constraints and was not applied", violations=violations
        )

    flips = _compute_flips(db, cycle, rows)
    descriptions = [
        f"{'Granting' if honored else 'Revoking'} {req.request_type.value} request #{req.id} for agent {req.agent_id}"
        for req, honored in flips
    ]
    if (require_reason or flips) and not reason:
        message = (
            "This edit overrides solver-decided request outcome(s); a reason is required"
            if flips
            else "A reason is required for a manual override"
        )
        raise RosterImportError(message, violations=descriptions, status_code=400)

    old_count = db.query(RosterAssignment).filter(RosterAssignment.roster_id == roster.id).count()
    db.query(RosterAssignment).filter(RosterAssignment.roster_id == roster.id).delete()
    for r in rows:
        db.add(
            RosterAssignment(
                roster_id=roster.id,
                agent_id=r.agent_id,
                date=r.date,
                shift_id=r.shift_id,
                skill_covered_id=r.skill_id,
                source=AssignmentSource.manual_override,
            )
        )

    for req, honored in flips:
        old_status = req.status
        if req.request_type in LEAVE_REQUEST_TYPES:
            try:
                sync_leave_balance(db, req, old_status, honored)
            except LeaveBalanceSyncError as exc:
                raise RosterImportError(exc.detail, status_code=400)
        req.status = RequestStatus.approved if honored else RequestStatus.denied
        req.denial_reason = None if honored else "manual roster override"
        db.add(
            AuditLog(
                actor_id=actor_id,
                action_type="request_outcome_overridden",
                target_type="weekly_request",
                target_id=req.id,
                old_value=old_status.value,
                new_value=req.status.value,
                reason=reason,
            )
        )

    db.add(
        AuditLog(
            actor_id=actor_id,
            action_type=action_type,
            target_type="roster",
            target_id=roster.id,
            old_value=f"{old_count} assignment(s)",
            new_value=f"{len(rows)} assignment(s)",
            reason=reason,
        )
    )

    db.commit()
    db.refresh(roster)
    return ImportResult(roster=roster, overridden_requests=descriptions)


def revalidate_and_apply_import(
    db: Session, roster_id: int, rows: list[ParsedAssignmentRow], reason: str | None, actor_id: int
) -> ImportResult:
    roster = db.query(Roster).filter(Roster.id == roster_id).first()
    if roster is None:
        raise RosterImportError("Roster not found", status_code=404)
    cycle = db.query(WeeklyCycle).filter(WeeklyCycle.id == roster.week_cycle_id).first()
    return apply_proposed_roster(
        db, roster, cycle, rows, reason, actor_id, action_type="roster_import_override", require_reason=False
    )


def _derive_skill_for_assignment(
    db: Session, agent_id: int, shift_id: int, target_date: date
) -> int:
    """Pick which of the agent's own skills they'll cover on a manual assignment.
    An override just needs a shift — the skill is inferred from the agent's skill
    set (they always cover something they're trained in), preferring a skill that
    actually has coverage demand on that shift's hours that day."""
    held = [link.skill_id for link in db.query(AgentSkill).filter(AgentSkill.agent_id == agent_id).all()]
    if not held:
        raise RosterImportError(
            f"Agent {agent_id} holds no skills, so they cannot be assigned to a shift", status_code=400
        )
    shift = db.query(ShiftTemplate).filter(ShiftTemplate.id == shift_id).first()
    if shift is None:
        raise RosterImportError(f"Shift template {shift_id} does not exist", status_code=400)

    coverage_reqs = (
        db.query(CoverageRequirement)
        .filter(
            CoverageRequirement.day_of_week == target_date.weekday(),
            CoverageRequirement.skill_id.in_(held),
        )
        .all()
    )
    matching = [
        req
        for req in coverage_reqs
        if shift_covers_slot(shift.start_time, shift.end_time, req.time_slot_start, req.time_slot_end)
    ]
    if matching:
        # cover the most-demanded matching skill
        matching.sort(key=lambda r: r.min_agents_required, reverse=True)
        return matching[0].skill_id
    return held[0]


def apply_manual_override(
    db: Session,
    roster_id: int,
    agent_id: int,
    target_date: date,
    shift_id: int | None,
    skill_id: int | None,
    reason: str,
    actor_id: int,
) -> ImportResult:
    """Edit a single (agent, date) assignment directly — set shift_id to assign
    the agent to that shift (the skill they cover is inferred from their own
    skill set unless skill_id is given), or leave shift_id None to unassign
    (give the agent that day off)."""
    roster = db.query(Roster).filter(Roster.id == roster_id).first()
    if roster is None:
        raise RosterImportError("Roster not found", status_code=404)
    cycle = db.query(WeeklyCycle).filter(WeeklyCycle.id == roster.week_cycle_id).first()

    current = db.query(RosterAssignment).filter(RosterAssignment.roster_id == roster.id).all()
    rows = [
        ParsedAssignmentRow(agent_id=a.agent_id, date=a.date, shift_id=a.shift_id, skill_id=a.skill_covered_id)
        for a in current
        if not (a.agent_id == agent_id and a.date == target_date)
    ]
    if shift_id is not None:
        resolved_skill = skill_id if skill_id is not None else _derive_skill_for_assignment(
            db, agent_id, shift_id, target_date
        )
        rows.append(ParsedAssignmentRow(agent_id=agent_id, date=target_date, shift_id=shift_id, skill_id=resolved_skill))

    return apply_proposed_roster(
        db, roster, cycle, rows, reason, actor_id, action_type="roster_manual_override", require_reason=True
    )
