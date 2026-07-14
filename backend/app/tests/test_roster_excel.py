from datetime import date, timedelta
from io import BytesIO

from openpyxl import Workbook, load_workbook


def _next_monday(weeks_ahead: int = 1) -> date:
    today = date.today()
    days_until_monday = (7 - today.weekday()) % 7 or 7
    return today + timedelta(days=days_until_monday + 7 * (weeks_ahead - 1))


def _create_cycle(client, manager_headers, weeks_ahead: int = 2) -> dict:
    monday = _next_monday(weeks_ahead)
    return client.post(
        "/api/weekly-cycles", json={"week_start_date": monday.isoformat()}, headers=manager_headers
    ).json()


def _create_skill(client, manager_headers, name="Prepaid Sales") -> dict:
    return client.post("/api/skills", json={"name": name}, headers=manager_headers).json()


def _create_shift(client, manager_headers, name="Day", start="09:00:00", end="17:00:00") -> dict:
    return client.post(
        "/api/shift-templates", json={"name": name, "start_time": start, "end_time": end}, headers=manager_headers
    ).json()


def _create_agent(client, manager_headers, name, skill_ids, shift_id) -> dict:
    return client.post(
        "/api/agents",
        json={"name": name, "default_shift_id": shift_id, "skill_ids": skill_ids},
        headers=manager_headers,
    ).json()


def _create_coverage(client, manager_headers, day_of_week, skill_id, min_required, start="09:00:00", end="17:00:00") -> dict:
    return client.post(
        "/api/coverage-requirements",
        json={
            "day_of_week": day_of_week,
            "time_slot_start": start,
            "time_slot_end": end,
            "skill_id": skill_id,
            "min_agents_required": min_required,
        },
        headers=manager_headers,
    ).json()


def _xlsx_bytes(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["agent_id", "date", "shift_id", "skill_id"])
    for r in rows:
        ws.append([r["agent_id"], r["date"], r["shift_id"], r["skill_id"]])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _setup_roster(client, manager_headers, num_agents=2, min_required=1):
    cycle = _create_cycle(client, manager_headers)
    skill = _create_skill(client, manager_headers)
    shift = _create_shift(client, manager_headers)
    agents = [_create_agent(client, manager_headers, f"Agent {i}", [skill["id"]], shift["id"]) for i in range(num_agents)]
    _create_coverage(client, manager_headers, day_of_week=0, skill_id=skill["id"], min_required=min_required)
    gen = client.post(f"/api/roster/generate?week_cycle_id={cycle['id']}", headers=manager_headers).json()
    return cycle, skill, shift, agents, gen["roster"]


def test_export_returns_xlsx_matching_assignments(client, manager_headers):
    cycle, skill, shift, agents, roster = _setup_roster(client, manager_headers, num_agents=1)

    response = client.get(f"/api/roster/{roster['id']}/export", headers=manager_headers)
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]

    wb = load_workbook(BytesIO(response.content))
    # First sheet is the human-friendly grid (agent rows x day columns).
    grid = wb.active
    assert grid.title == "Roster"
    assert grid.cell(row=1, column=1).value == "Agent"
    assert str(grid.cell(row=1, column=2).value).startswith("Monday")
    assert grid.cell(row=2, column=1).value == agents[0]["name"]

    # Second sheet keeps the machine-readable long format for re-upload.
    data = wb["Data"]
    rows = list(data.iter_rows(values_only=True))
    assert rows[0][:4] == ("agent_id", "agent_name", "date", "shift_id")
    assert len(rows) >= 2  # header + at least 1 assignment


def test_import_rejects_double_booking(client, manager_headers):
    cycle, skill, shift, agents, roster = _setup_roster(client, manager_headers, num_agents=1)
    agent_id = agents[0]["id"]
    monday = cycle["week_start_date"]

    file_bytes = _xlsx_bytes(
        [
            {"agent_id": agent_id, "date": monday, "shift_id": shift["id"], "skill_id": skill["id"]},
            {"agent_id": agent_id, "date": monday, "shift_id": shift["id"], "skill_id": skill["id"]},
        ]
    )
    response = client.post(
        f"/api/roster/{roster['id']}/import",
        headers=manager_headers,
        files={"file": ("roster.xlsx", file_bytes, "application/octet-stream")},
    )
    assert response.status_code == 422
    assert "double-booked" in str(response.json()["detail"]["violations"])


def test_import_rejects_coverage_shortfall(client, manager_headers):
    cycle, skill, shift, agents, roster = _setup_roster(client, manager_headers, num_agents=2, min_required=2)
    monday = cycle["week_start_date"]

    # Only 1 of the 2 required agents scheduled -> coverage shortfall
    file_bytes = _xlsx_bytes(
        [{"agent_id": agents[0]["id"], "date": monday, "shift_id": shift["id"], "skill_id": skill["id"]}]
    )
    response = client.post(
        f"/api/roster/{roster['id']}/import",
        headers=manager_headers,
        files={"file": ("roster.xlsx", file_bytes, "application/octet-stream")},
    )
    assert response.status_code == 422
    assert "Coverage shortfall" in str(response.json()["detail"]["violations"])


def test_import_rejects_unknown_agent(client, manager_headers):
    cycle, skill, shift, agents, roster = _setup_roster(client, manager_headers, num_agents=1)
    monday = cycle["week_start_date"]

    file_bytes = _xlsx_bytes([{"agent_id": 999999, "date": monday, "shift_id": shift["id"], "skill_id": skill["id"]}])
    response = client.post(
        f"/api/roster/{roster['id']}/import",
        headers=manager_headers,
        files={"file": ("roster.xlsx", file_bytes, "application/octet-stream")},
    )
    assert response.status_code == 422
    assert "does not exist" in str(response.json()["detail"]["violations"])


def test_import_valid_edit_applies_cleanly(client, manager_headers):
    cycle, skill, shift, agents, roster = _setup_roster(client, manager_headers, num_agents=2, min_required=1)
    monday = cycle["week_start_date"]

    # Swap which of the two interchangeable agents covers Monday -> no flips, no reason needed
    file_bytes = _xlsx_bytes(
        [{"agent_id": agents[1]["id"], "date": monday, "shift_id": shift["id"], "skill_id": skill["id"]}]
    )
    response = client.post(
        f"/api/roster/{roster['id']}/import",
        headers=manager_headers,
        files={"file": ("roster.xlsx", file_bytes, "application/octet-stream")},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["overridden_requests"] == []
    assert len(body["assignments"]) == 1
    assert body["assignments"][0]["agent_id"] == agents[1]["id"]
    assert body["assignments"][0]["source"] == "manual_override"


def test_import_blocks_working_on_approved_leave(client, manager_headers, agent_headers, agent_record):
    cycle = _create_cycle(client, manager_headers)
    skill = _create_skill(client, manager_headers)
    shift = _create_shift(client, manager_headers)
    client.patch(
        f"/api/agents/{agent_record.id}",
        json={"default_shift_id": shift["id"], "skill_ids": [skill["id"]]},
        headers=manager_headers,
    )
    client.post(
        "/api/leave-balance",
        json={"agent_id": agent_record.id, "year": date.today().year, "total_leave_days_allotted": 5},
        headers=manager_headers,
    )
    monday = cycle["week_start_date"]
    client.post(
        "/api/requests",
        json={"week_cycle_id": cycle["id"], "request_type": "leave_full", "requested_start_date": monday},
        headers=agent_headers,
    )
    # No coverage requirement -> leave is honored (agent off, balance decremented)
    gen = client.post(f"/api/roster/generate?week_cycle_id={cycle['id']}", headers=manager_headers).json()
    roster = gen["roster"]

    # Try to schedule the agent to work anyway on their approved leave day
    file_bytes = _xlsx_bytes(
        [{"agent_id": agent_record.id, "date": monday, "shift_id": shift["id"], "skill_id": skill["id"]}]
    )
    response = client.post(
        f"/api/roster/{roster['id']}/import",
        headers=manager_headers,
        files={"file": ("roster.xlsx", file_bytes, "application/octet-stream")},
    )
    assert response.status_code == 422
    assert "approved leave" in str(response.json()["detail"]["violations"])


def test_import_reversing_off_day_requires_reason(client, manager_headers, agent_headers, agent_record):
    cycle = _create_cycle(client, manager_headers)
    skill = _create_skill(client, manager_headers)
    shift = _create_shift(client, manager_headers)
    client.patch(
        f"/api/agents/{agent_record.id}",
        json={"default_shift_id": shift["id"], "skill_ids": [skill["id"]]},
        headers=manager_headers,
    )
    second_agent = _create_agent(client, manager_headers, "Backup", [skill["id"]], shift["id"])
    _create_coverage(client, manager_headers, day_of_week=0, skill_id=skill["id"], min_required=1)
    monday = cycle["week_start_date"]
    req = client.post(
        "/api/requests",
        json={"week_cycle_id": cycle["id"], "request_type": "off_day", "requested_start_date": monday},
        headers=agent_headers,
    ).json()

    # Solver honors the off-day (second agent covers instead)
    gen = client.post(f"/api/roster/generate?week_cycle_id={cycle['id']}", headers=manager_headers).json()
    roster = gen["roster"]
    assert gen["conflicts"] == []

    # Manager tries to override: schedule the requester to work anyway, without a reason
    file_bytes = _xlsx_bytes(
        [{"agent_id": agent_record.id, "date": monday, "shift_id": shift["id"], "skill_id": skill["id"]}]
    )
    response = client.post(
        f"/api/roster/{roster['id']}/import",
        headers=manager_headers,
        files={"file": ("roster.xlsx", file_bytes, "application/octet-stream")},
    )
    assert response.status_code == 400
    assert "reason is required" in response.json()["detail"]["message"]

    # Now with a reason: should apply and flip the request to denied
    response = client.post(
        f"/api/roster/{roster['id']}/import",
        headers=manager_headers,
        data={"reason": "Emergency coverage need"},
        files={"file": ("roster.xlsx", file_bytes, "application/octet-stream")},
    )
    assert response.status_code == 200
    body = response.json()
    assert len(body["overridden_requests"]) == 1

    updated_request = client.get("/api/requests/mine", headers=agent_headers).json()[0]
    assert updated_request["id"] == req["id"]
    assert updated_request["status"] == "denied"
    assert updated_request["denial_reason"] == "manual roster override"


def test_import_blocked_on_locked_cycle(client, manager_headers):
    cycle, skill, shift, agents, roster = _setup_roster(client, manager_headers, num_agents=1)
    client.post(f"/api/roster/{roster['id']}/publish", headers=manager_headers)
    client.post(f"/api/roster/{roster['id']}/lock", headers=manager_headers)

    file_bytes = _xlsx_bytes(
        [{"agent_id": agents[0]["id"], "date": cycle["week_start_date"], "shift_id": shift["id"], "skill_id": skill["id"]}]
    )
    response = client.post(
        f"/api/roster/{roster['id']}/import",
        headers=manager_headers,
        files={"file": ("roster.xlsx", file_bytes, "application/octet-stream")},
    )
    assert response.status_code == 400


def test_import_requires_manager(client, agent_headers):
    file_bytes = _xlsx_bytes([])
    response = client.post(
        "/api/roster/1/import",
        headers=agent_headers,
        files={"file": ("roster.xlsx", file_bytes, "application/octet-stream")},
    )
    assert response.status_code == 403
